import pandas as pd
from datetime import datetime, time, timedelta
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _find_column(df, *keyword_groups):
    for keywords in keyword_groups:
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in keywords):
                return col
    return None


def _detect_columns(df):
    time_col = _find_column(df,
        ["punchcardtime", "punch_card_time"],
        ["punchcard", "punch_card"],
        ["timestamp", "datetime"],
        ["time"],
    )
    entry_exit_col = _find_column(df,
        ["warehouseentry", "exitidentifier"],
        ["identifier", "entry/exit", "entryexit"],
        ["entry", "exit"],
        ["checkin", "check-in", "check_in"],
    )
    employee_col = _find_column(df,
        ["employeename", "employee_name"],
        ["employee"],
        ["name", "worker", "staff"],
    )
    return time_col, entry_exit_col, employee_col


def _validate_columns(df, time_col, entry_exit_col, employee_col):
    missing = []
    if not time_col:       missing.append("PunchCard Time")
    if not entry_exit_col: missing.append("Entry/Exit Identifier")
    if not employee_col:   missing.append("Employee Name")
    if missing:
        available = ", ".join(str(c) for c in df.columns)
        raise ValueError(
            f"Could not detect column(s): {', '.join(missing)}\n"
            f"Available columns: {available}"
        )


def _is_checkin(value):
    val = str(value).lower().strip()
    return "check-in" in val or "checkin" in val or val == "0"


def _is_checkout(value):
    val = str(value).lower().strip()
    return "check-out" in val or "checkout" in val or val == "1"


def _fmt_td(td):
    total_sec = int(td.total_seconds())
    sign = "-" if total_sec < 0 else ""
    total_sec = abs(total_sec)
    h, rem = divmod(total_sec, 3600)
    m, s = divmod(rem, 60)
    return f"{sign}{h:02d}:{m:02d}:{s:02d}"


def _fmt_hours(decimal_hours):
    """Convert decimal hours to 'Xh Ym' string (e.g. 2.5 → '2h 30m')."""
    if decimal_hours <= 0:
        return "—"
    total_min = round(decimal_hours * 60)
    h, m = divmod(total_min, 60)
    return f"{h}h {m:02d}m"


def detect_issues(input_file):
    """Quick parse — returns (issues_list, sorted_employees_list)."""
    df = pd.read_excel(input_file)
    time_col, entry_exit_col, employee_col = _detect_columns(df)
    _validate_columns(df, time_col, entry_exit_col, employee_col)

    df[time_col] = pd.to_datetime(df[time_col])
    df["_Date"] = df[time_col].dt.date

    issues = []
    for (day, emp), group in df.groupby(["_Date", employee_col]):
        has_in  = group[entry_exit_col].apply(_is_checkin).any()
        has_out = group[entry_exit_col].apply(_is_checkout).any()
        if not has_in:
            issues.append({"Date": str(day), "Employee": emp, "Problem": "Missing Check-in"})
        elif not has_out:
            issues.append({"Date": str(day), "Employee": emp, "Problem": "Missing Check-out"})

    employees = sorted(df[employee_col].dropna().unique().tolist())
    return issues, employees


def generate_report(
    input_file,
    am_start_hour=5, am_start_min=30,
    pm_start_hour=13, pm_start_min=0,
    break_hours=1.0,
    manual_entries=None,
):
    df = pd.read_excel(input_file)
    time_col, entry_exit_col, employee_col = _detect_columns(df)
    _validate_columns(df, time_col, entry_exit_col, employee_col)

    # Merge manual entries into the dataframe before processing
    if manual_entries:
        manual_rows = [
            {
                time_col:       pd.Timestamp(e["datetime"]),
                entry_exit_col: e["type"],
                employee_col:   e["employee"],
            }
            for e in manual_entries
        ]
        df = pd.concat([df, pd.DataFrame(manual_rows)], ignore_index=True)

    df[time_col] = pd.to_datetime(df[time_col])
    df["_Date"] = df[time_col].dt.date

    rows, skipped = [], []

    for (day, emp), group in df.groupby(["_Date", employee_col]):
        group = group.sort_values(time_col)

        checkins  = group[group[entry_exit_col].apply(_is_checkin)]
        checkouts = group[group[entry_exit_col].apply(_is_checkout)]

        if checkins.empty:
            skipped.append(f"{day}  /  {emp}  →  no check-in found")
            continue
        if checkouts.empty:
            skipped.append(f"{day}  /  {emp}  →  no check-out found")
            continue

        first_in = checkins[time_col].min()
        last_out = checkouts[time_col].max()

        if first_in.hour < 12:
            shift    = "AM"
            start_dt = datetime.combine(day, time(am_start_hour, am_start_min))
        else:
            shift    = "PM"
            start_dt = datetime.combine(day, time(pm_start_hour, pm_start_min))

        total_delta     = last_out - start_dt - timedelta(hours=break_hours)
        total_hours_val = total_delta.total_seconds() / 3600
        break_label     = f"{int(break_hours)}h" if break_hours == int(break_hours) else f"{break_hours}h"

        rows.append({
            "Date":       day,
            "Employee":   emp,
            "Shift":      shift,
            "Start Time": start_dt.strftime("%H:%M"),
            "First Clock In":  first_in.strftime("%H:%M:%S"),
            "Last Clock Out":  last_out.strftime("%H:%M:%S"),
            f"Total Time (-{break_label} break)": _fmt_td(total_delta),
            "_total_hours": round(total_hours_val, 4),
        })

    if not rows:
        raise ValueError(
            "No valid records found.\n"
            "Make sure the file contains Check-in and Check-out entries."
        )

    daily = (
        pd.DataFrame(rows)
        .sort_values(["Date", "Employee"])
        .reset_index(drop=True)
    )

    # Weekly summary
    dates = pd.to_datetime(daily["Date"])
    iso   = dates.dt.isocalendar()
    daily["_year"]       = iso.year.astype(int)
    daily["_week"]       = iso.week.astype(int)
    daily["_week_start"] = dates - pd.to_timedelta(dates.dt.dayofweek, unit="D")

    weekly = (
        daily
        .groupby(["_year", "_week", "_week_start", "Employee"])["_total_hours"]
        .sum()
        .reset_index()
    )
    weekly.columns = ["Year", "Week", "Week Start", "Employee", "Total Hours"]
    weekly["Total Hours"]    = weekly["Total Hours"].round(2)
    weekly["Regular Hours"]  = weekly["Total Hours"].apply(lambda x: round(min(x, 40), 2))
    weekly["Overtime Hours"] = weekly["Total Hours"].apply(lambda x: _fmt_hours(max(0.0, x - 40)))
    weekly["Week Start"]     = weekly["Week Start"].dt.date
    weekly = weekly.sort_values(["Year", "Week", "Employee"]).reset_index(drop=True)

    break_col     = f"Total Time (-{break_label} break)"
    daily_export  = daily[["Date", "Employee", "Shift", "Start Time",
                            "First Clock In", "Last Clock Out", break_col]].copy()
    weekly_export = weekly[["Week Start", "Year", "Week", "Employee",
                             "Total Hours", "Regular Hours", "Overtime Hours"]].copy()

    # ── Monthly summary ───────────────────────────────────────────────────────
    dates_m = pd.to_datetime(daily["Date"])
    daily["_month_num"]  = dates_m.dt.month
    daily["_month_name"] = dates_m.dt.strftime("%B %Y")

    monthly = (
        daily
        .groupby(["_year", "_month_num", "_month_name", "Employee"])
        .agg(days_worked=("_total_hours", "count"), total_hours=("_total_hours", "sum"))
        .reset_index()
    )
    monthly["total_hours"]    = monthly["total_hours"].round(2)
    monthly["regular_hours"]  = monthly["total_hours"].apply(lambda x: round(min(x, 160), 2))
    monthly["overtime_hours"] = monthly["total_hours"].apply(lambda x: _fmt_hours(max(0.0, x - 160)))
    monthly = monthly.sort_values(["_year", "_month_num", "Employee"]).reset_index(drop=True)

    monthly_export = monthly.rename(columns={
        "_month_name":   "Month",
        "_year":         "Year",
        "Employee":      "Employee",
        "days_worked":   "Days Worked",
        "total_hours":   "Total Hours",
        "regular_hours": "Regular Hours",
        "overtime_hours":"Overtime Hours",
    })[["Month", "Year", "Employee", "Days Worked", "Total Hours", "Regular Hours", "Overtime Hours"]].copy()

    return daily_export, weekly_export, monthly_export, daily["_total_hours"].sum(), skipped


# ── Excel export ──────────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT   = Font(color="FFFFFF", bold=True)
_OVERTIME_FILL = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
_CENTER        = Alignment(horizontal="center", vertical="center")


def _style_sheet(ws, df, overtime_col=None):
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER

    if overtime_col:
        ot_idx = df.columns.get_loc(overtime_col) + 1
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=ot_idx).value
            has_ot = val is not None and val != "—" and val != "" and val != 0
            if has_ot:
                for ci in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=ci).fill = _OVERTIME_FILL

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def create_excel(daily_df, weekly_df, monthly_df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        daily_df.to_excel(writer,   sheet_name="Daily Report",    index=False)
        weekly_df.to_excel(writer,  sheet_name="Weekly Summary",  index=False)
        monthly_df.to_excel(writer, sheet_name="Monthly Summary", index=False)
        _style_sheet(writer.sheets["Daily Report"],    daily_df)
        _style_sheet(writer.sheets["Weekly Summary"],  weekly_df,  overtime_col="Overtime Hours")
        _style_sheet(writer.sheets["Monthly Summary"], monthly_df, overtime_col="Overtime Hours")
    buf.seek(0)
    return buf
